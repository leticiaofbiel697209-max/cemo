from __future__ import annotations

import json
import re
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook
from sqlalchemy import select

from database.models import Estoque, ImportacaoPlanilha, Insumo
from services.auditoria_service import registrar
from services.estoque_service import atualizar_calculos
from utils.helpers import parse_decimal

FIELD_ALIASES = {
    "codigo": ["codigo", "código", "cod", "item", "catmat"],
    "descricao": ["descricao", "descrição", "insumo", "material", "produto", "objeto"],
    "categoria": ["categoria", "grupo", "classe"],
    "subcategoria": ["subcategoria", "sub grupo", "subgrupo"],
    "unidade_medida": ["unidade", "und", "unid", "u.m.", "um"],
    "fabricante": ["fabricante"],
    "marca": ["marca"],
    "fornecedor_principal": ["fornecedor", "empresa"],
    "observacoes": ["observacao", "observação", "obs"],
    "quantidade_atual": ["estoque", "saldo", "quantidade atual", "qtd atual"],
    "estoque_minimo": ["estoque mínimo", "minimo", "mínimo"],
    "consumo_medio_mensal": ["consumo médio", "consumo mensal", "cmm"],
}


def normalize(text: str) -> str:
    text = str(text or "").strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def detect_header(raw: pd.DataFrame) -> int:
    best_idx, best_score = 0, -1
    words = {alias for values in FIELD_ALIASES.values() for alias in values}
    for idx, row in raw.head(20).iterrows():
        values = [normalize(v) for v in row.tolist() if str(v).strip() and str(v) != "nan"]
        score = sum(any(alias in value for alias in words) for value in values)
        if score > best_score:
            best_idx, best_score = idx, score
    return int(best_idx)


def read_workbook(file_bytes: bytes) -> dict[str, pd.DataFrame]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True)
    for ws in wb.worksheets:
        for merged in list(ws.merged_cells.ranges):
            value = ws.cell(merged.min_row, merged.min_col).value
            ws.unmerge_cells(str(merged))
            for row in ws.iter_rows(merged.min_row, merged.max_row, merged.min_col, merged.max_col):
                for cell in row:
                    cell.value = value
    output: dict[str, pd.DataFrame] = {}
    for ws in wb.worksheets:
        rows = list(ws.values)
        raw = pd.DataFrame(rows).dropna(how="all")
        if raw.empty:
            continue
        header_idx = detect_header(raw)
        header = [str(c).strip() if c else f"coluna_{i+1}" for i, c in enumerate(raw.iloc[header_idx].tolist())]
        df = raw.iloc[header_idx + 1 :].copy()
        df.columns = header
        df = df.dropna(how="all").reset_index(drop=True)
        output[ws.title] = df
    return output


def suggest_mapping(columns: list[str]) -> dict[str, str | None]:
    mapping: dict[str, str | None] = {}
    normalized = {col: normalize(col) for col in columns}
    for field, aliases in FIELD_ALIASES.items():
        mapping[field] = next((col for col, norm in normalized.items() if any(alias in norm for alias in aliases)), None)
    return mapping


def preview(file_bytes: bytes) -> tuple[dict[str, pd.DataFrame], dict[str, dict[str, str | None]]]:
    sheets = read_workbook(file_bytes)
    return sheets, {name: suggest_mapping(list(df.columns)) for name, df in sheets.items()}


def import_sheet(session, filename: str, sheets: dict[str, pd.DataFrame], mappings: dict[str, dict[str, str | None]], usuario_id: int | None = None) -> ImportacaoPlanilha:
    imported = ignored = 0
    report = []
    for sheet_name, df in sheets.items():
        mapping = mappings.get(sheet_name, {})
        if not mapping.get("descricao"):
            ignored += len(df)
            report.append({"aba": sheet_name, "status": "ignorada", "motivo": "Sem coluna de descrição mapeada."})
            continue
        for _, row in df.iterrows():
            descricao = str(row.get(mapping["descricao"], "")).strip()
            if not descricao or descricao.lower() == "nan":
                ignored += 1
                continue
            codigo_col = mapping.get("codigo")
            codigo = str(row.get(codigo_col, "")).strip() if codigo_col else None
            codigo = codigo if codigo and codigo.lower() != "nan" else None
            existing = None
            if codigo:
                existing = session.execute(select(Insumo).where(Insumo.codigo == codigo)).scalar_one_or_none()
            if not existing:
                existing = session.execute(select(Insumo).where(Insumo.descricao == descricao)).scalar_one_or_none()
            origem = {col: str(row.get(col, "")) for col in df.columns if col not in mapping.values()}
            insumo = existing or Insumo(descricao=descricao, codigo=codigo)
            for field in ["categoria", "subcategoria", "unidade_medida", "fabricante", "marca", "fornecedor_principal", "observacoes"]:
                col = mapping.get(field)
                if col and pd.notna(row.get(col)):
                    setattr(insumo, field, str(row.get(col)).strip())
            insumo.dados_origem = json.dumps({"aba": sheet_name, "campos_nao_mapeados": origem}, ensure_ascii=False)
            session.add(insumo)
            session.flush()
            estoque = insumo.estoque or Estoque(insumo_id=insumo.id)
            for field in ["quantidade_atual", "estoque_minimo", "consumo_medio_mensal"]:
                col = mapping.get(field)
                value = parse_decimal(row.get(col)) if col else None
                if value is not None:
                    setattr(estoque, field, value)
            atualizar_calculos(estoque)
            session.add(estoque)
            imported += 0 if existing else 1
    imp = ImportacaoPlanilha(
        nome_arquivo=filename,
        abas_processadas=len(sheets),
        registros_importados=imported,
        registros_ignorados=ignored,
        relatorio=json.dumps(report, ensure_ascii=False),
        usuario_id=usuario_id,
    )
    session.add(imp)
    registrar(session, usuario_id, "importacao", filename, "importação", depois={"importados": imported, "ignorados": ignored})
    return imp
